import os
from datetime import datetime
import pandas as pd
import pytest
from openpyxl import load_workbook
import logging
from configparser import ConfigParser



###### create config parser ########
# create an instance of configparser
config = ConfigParser()
# reading config.ini file
config.read("configuration/config.ini")

# Create a logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

# Directory for log files
log_path = config["logs"]["log_file_path"]
os.makedirs(log_path, exist_ok=True)  # Create 'Logs' directory if it doesn't exist

# Generate current timestamp
current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

# Log file name with timestamp
log_file = os.path.join(log_path, f'csv_gen_execution_{current_datetime}.log')

# Create a handler for writing log messages to a file
file_handler = logging.FileHandler(log_file)
file_handler.setLevel(logging.DEBUG)

# Create a formatter for the log messages
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

# Add the file handler to the logger
logger.addHandler(file_handler)

# Example usage:
logger.debug('Debug message')
logger.info('Info message')
logger.warning('Warning message')
logger.error('Error message')
logger.critical('Critical message')


@pytest.fixture
def get_the_excel_sheet_names_with_prefix():
    logger.info("Fetching Excel sheet names with prefix 'TC_'")
    excel_path = config["files"]["excel_input_path"]
    workbook = load_workbook(excel_path, read_only=True)
    sheet_names = [sheet for sheet in workbook.sheetnames if sheet.startswith('TC_')]
    logger.debug(f"Found sheet names: {sheet_names}")

    if not sheet_names:
        logger.error("No sheets found with prefix 'TC_', Please add your Test Cases accordingly.")
        raise RuntimeError("No sheets found with prefix 'TC_', Please add your Test Cases accordingly.")

    return sheet_names


def exclude_sheets(sheet_names):
    logger.info("Excluding sheets based on user preference")
    exclude_sheets = "yes"
    if exclude_sheets == config["preference"]["exclude"]:
        exclude_list = config["preference"]["exclude_sheets"].strip().split(',')
        exclude_list = [sheet.strip() for sheet in exclude_list]
        sheet_names = [sheet for sheet in sheet_names if sheet not in exclude_list]
        logger.debug(f"Excluded sheets: {exclude_list}")
    else:
        logger.debug("No sheets excluded as per user choice")
    return sheet_names


@pytest.fixture
def exclude_sheets_(get_the_excel_sheet_names_with_prefix):
    sheet_names = get_the_excel_sheet_names_with_prefix
    return exclude_sheets(sheet_names)


@pytest.fixture
def create_csv_from_sheets(exclude_sheets_):
    logger.info("Creating CSV from selected sheets")
    excel_path = config["files"]["excel_input_path"]
    df_combined = pd.DataFrame()

    for sheet_name in exclude_sheets_:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        df_combined = pd.concat([df_combined, df], ignore_index=True)

    csv_path = config["files"]["csv_output_path"]
    df_combined.to_csv(csv_path, index=False)
    logger.info(f"CSV created and saved to {csv_path}")

    return csv_path